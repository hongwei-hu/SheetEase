"""导表流程模块测试。"""

import json
from pathlib import Path

import openpyxl
import pytest

from ExcelExportTool.core import export_process
from ExcelExportTool.exceptions import ExportError, SheetNameConflictError
from ExcelExportTool.generation.enum_registry import get_enum_registry, reset_enum_registry


def test_process_excel_file_sheet_name_conflict_raises():
    wb = openpyxl.Workbook()
    wb.active.title = "SameSheet"

    with pytest.raises(SheetNameConflictError):
        export_process.process_excel_file(
            excel_path=Path("B.xlsx"),
            file_sheet_map={"A.xlsx": "SameSheet"},
            output_client_folder=None,
            output_project_folder=None,
            csfile_output_folder=None,
            enum_output_folder=None,
            workbook=wb,
        )


def test_process_excel_file_calls_generate_methods(monkeypatch, tmp_path):
    class DummyWSData:
        def __init__(self, ws):
            self.name = ws.title
            self.generated_json = 0
            self.generated_script = 0

        def generate_json(self, output_folder):
            self.generated_json += 1

        def generate_script(self, output_folder):
            self.generated_script += 1

    wb = openpyxl.Workbook()
    wb.active.title = "Item"

    monkeypatch.setattr(export_process, "WorksheetData", DummyWSData)

    result = export_process.process_excel_file(
        excel_path=Path("Item.xlsx"),
        file_sheet_map={},
        output_client_folder=str(tmp_path / "client"),
        output_project_folder=str(tmp_path / "project"),
        csfile_output_folder=str(tmp_path / "cs"),
        enum_output_folder=None,
        workbook=wb,
    )

    assert result is not None
    assert result.generated_json == 2
    assert result.generated_script == 1


def test_batch_excel_to_json_warns_when_no_files(monkeypatch, tmp_path):
    warned = []
    monkeypatch.setattr(export_process, "log_warn", lambda msg, **kwargs: warned.append(msg))

    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=None,
        auto_cleanup=False,
    )

    assert any("未找到 .xlsx 文件" in m for m in warned)


def test_batch_excel_to_json_output_not_writable_raises(monkeypatch, tmp_path):
    out_dir = tmp_path / "out"
    out_dir.mkdir(parents=True, exist_ok=True)

    original_access = export_process.os.access

    def fake_access(path, mode):
        if str(path) == str(out_dir):
            return False
        return original_access(path, mode)

    monkeypatch.setattr(export_process.os, "access", fake_access)

    with pytest.raises(ExportError, match="输出目录不可写"):
        export_process.batch_excel_to_json(
            source_folder=str(tmp_path),
            output_client_folder=None,
            output_project_folder=str(out_dir),
            csfile_output_folder=None,
            enum_output_folder=None,
            auto_cleanup=False,
        )


def test_cleanup_files_ignores_sheetease_metadata(monkeypatch, tmp_path):
    metadata_file = tmp_path / ".sheetease" / "stable_enum_values.json"
    metadata_file.parent.mkdir(parents=True)
    metadata_file.write_text("{}", encoding="utf-8")

    def fail_confirm(*args, **kwargs):
        raise AssertionError("cleanup should not ask to delete SheetEase metadata")

    monkeypatch.setattr(export_process, "user_confirm", fail_confirm)

    export_process.cleanup_files([str(tmp_path)])

    assert metadata_file.exists()


def _build_minimal_workbook_for_enum_collection(path: Path, type_col: int) -> None:
    """创建仅用于第一阶段枚举收集的最小工作簿。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = path.stem
    # 第3行：类型；第7行：第一条数据
    ws.cell(row=3, column=type_col, value="string")
    ws.cell(row=5, column=type_col, value="weapon_template")
    ws.cell(row=7, column=type_col, value="SwordTemplate")
    wb.save(path)
    wb.close()


def _build_workbook_for_keys_enum_collection(path: Path, type_col: int, keys: list[str]) -> None:
    """创建用于字符串主键枚举收集的工作簿，支持多条 key。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = path.stem
    ws.cell(row=3, column=type_col, value="string")
    ws.cell(row=5, column=type_col, value="keyName")
    for offset, key in enumerate(keys):
        ws.cell(row=7 + offset, column=type_col, value=key)
    wb.save(path)
    wb.close()


def _build_workbook_with_enum_sheet(path: Path, enum_rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = path.stem
    enum_ws = wb.create_sheet("Enum-Quality")
    enum_ws.append(["name", "value", "remark"])
    for row in enum_rows:
        enum_ws.append(list(row))
    wb.save(path)
    wb.close()


def _write_generated_enum(path: Path, enum_name: str, members: list[tuple[str, int]]) -> None:
    lines = [
        "namespace Data.TableScript",
        "{",
        f"    public enum {enum_name}",
        "    {",
    ]
    for name, value in members:
        lines.append(f"        {name} = {value},")
    lines.extend(["    }", "}"])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def test_batch_excel_to_json_collects_keys_enum_from_column_b(monkeypatch, tmp_path):
    # 文件名首字母大写才会参与导表
    xlsx = tmp_path / "WeaponTemplate.xlsx"
    _build_minimal_workbook_for_enum_collection(xlsx, type_col=2)  # B列

    # 避免进入第二阶段复杂解析，专注验证第一阶段枚举收集
    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=None,
        auto_cleanup=False,
    )

    reg = get_enum_registry()
    assert reg.has_enum("WeaponTemplateKeys")
    assert reg.get_enum_value("WeaponTemplateKeys", "SwordTemplate") == 0


def test_batch_excel_to_json_collects_keys_enum_from_column_a_fallback(monkeypatch, tmp_path):
    xlsx = tmp_path / "LegacyTemplate.xlsx"
    _build_minimal_workbook_for_enum_collection(xlsx, type_col=1)  # A列（旧模板回退）

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=None,
        auto_cleanup=False,
    )

    reg = get_enum_registry()
    assert reg.has_enum("LegacyTemplateKeys")
    assert reg.get_enum_value("LegacyTemplateKeys", "SwordTemplate") == 0


def test_batch_excel_to_json_preserves_auto_key_values_when_inserted_between_runs(monkeypatch, tmp_path):
    xlsx = tmp_path / "WeaponTemplate.xlsx"
    _build_workbook_for_keys_enum_collection(xlsx, type_col=2, keys=["SwordTemplate", "BowTemplate"])

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)
    project_out = tmp_path / "project_out"

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=str(project_out),
        csfile_output_folder=None,
        enum_output_folder=None,
        auto_cleanup=False,
    )

    manifest_path = project_out / ".sheetease" / "stable_enum_values.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["WeaponTemplateKeys"] == {
        "SwordTemplate": 0,
        "BowTemplate": 1,
    }

    _build_workbook_for_keys_enum_collection(
        xlsx,
        type_col=2,
        keys=["SwordTemplate", "AxeTemplate", "BowTemplate"],
    )

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=str(project_out),
        csfile_output_folder=None,
        enum_output_folder=None,
        auto_cleanup=False,
    )

    reg = get_enum_registry()
    assert reg.get_enum_value("WeaponTemplateKeys", "SwordTemplate") == 0
    assert reg.get_enum_value("WeaponTemplateKeys", "BowTemplate") == 1
    assert reg.get_enum_value("WeaponTemplateKeys", "AxeTemplate") == 2


def test_batch_excel_to_json_uses_client_manifest_when_project_output_missing(monkeypatch, tmp_path):
    xlsx = tmp_path / "WeaponTemplate.xlsx"
    _build_workbook_for_keys_enum_collection(xlsx, type_col=2, keys=["SwordTemplate"])

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)
    client_out = tmp_path / "client_out"

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=str(client_out),
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=None,
        auto_cleanup=False,
    )

    manifest = json.loads((client_out / ".sheetease" / "stable_enum_values.json").read_text(encoding="utf-8"))
    assert manifest["WeaponTemplateKeys"]["SwordTemplate"] == 0
    assert not (tmp_path / ".stable_enum_values.json").exists()


def test_batch_excel_to_json_bootstraps_auto_key_values_from_existing_cs_output(monkeypatch, tmp_path):
    xlsx = tmp_path / "WeaponTemplate.xlsx"
    _build_workbook_for_keys_enum_collection(
        xlsx,
        type_col=2,
        keys=["SwordTemplate", "AxeTemplate", "BowTemplate"],
    )
    cs_out = tmp_path / "cs_out"
    project_out = tmp_path / "project_out"
    _write_generated_enum(
        cs_out / "WeaponTemplateKeys.cs",
        "WeaponTemplateKeys",
        [("SwordTemplate", 0), ("BowTemplate", 5)],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=str(project_out),
        csfile_output_folder=str(cs_out),
        enum_output_folder=None,
        auto_cleanup=False,
    )

    reg = get_enum_registry()
    assert reg.get_enum_value("WeaponTemplateKeys", "SwordTemplate") == 0
    assert reg.get_enum_value("WeaponTemplateKeys", "BowTemplate") == 5
    assert reg.get_enum_value("WeaponTemplateKeys", "AxeTemplate") == 6

    manifest = json.loads((project_out / ".sheetease" / "stable_enum_values.json").read_text(encoding="utf-8"))
    assert manifest["WeaponTemplateKeys"]["BowTemplate"] == 5


def test_batch_excel_to_json_does_not_export_auto_keys_to_enum_folder(monkeypatch, tmp_path):
    xlsx = tmp_path / "CombatAttribute.xlsx"
    _build_minimal_workbook_for_enum_collection(xlsx, type_col=2)  # B列 string 主键

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    exported_enum_names = []

    def _fake_generate_enum_file(enum_type_name, enum_names, enum_values, remarks, name_space, output_folder):
        exported_enum_names.append(enum_type_name)

    monkeypatch.setattr(export_process, "generate_enum_file", _fake_generate_enum_file)

    enum_out = tmp_path / "enum_out"
    enum_out.mkdir(parents=True, exist_ok=True)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=str(enum_out),
        auto_cleanup=False,
    )

    # 自动 Keys 枚举只注册，不写入 enum_output_folder
    assert "CombatAttributeKeys" not in exported_enum_names

    # 但运行时注册表仍可用（供 Enum(...) 字段校验与转换）
    reg = get_enum_registry()
    assert reg.has_enum("CombatAttributeKeys")


def test_batch_excel_to_json_exports_enum_sheet_items_without_explicit_values(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", None, "common"),
            ("Rare", None, "rare"),
        ],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    exported = {}

    def _fake_generate_enum_file(enum_type_name, enum_names, enum_values, remarks, name_space, output_folder):
        exported["enum_type_name"] = enum_type_name
        exported["enum_names"] = enum_names
        exported["enum_values"] = enum_values
        exported["remarks"] = remarks

    monkeypatch.setattr(export_process, "generate_enum_file", _fake_generate_enum_file)

    enum_out = tmp_path / "enum_out"
    enum_out.mkdir(parents=True, exist_ok=True)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=str(enum_out),
        auto_cleanup=False,
    )

    assert exported["enum_type_name"] == "Quality"
    assert exported["enum_names"] == ["Common", "Rare"]
    assert exported["enum_values"] == [0, 1]

    reg = get_enum_registry()
    assert reg.get_enum_value("Quality", "Common") == 0
    assert reg.get_enum_value("Quality", "Rare") == 1


def test_batch_excel_to_json_bootstraps_enum_sheet_values_from_existing_enum_output(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", None, "common"),
            ("Epic", None, "epic"),
            ("Rare", None, "rare"),
        ],
    )
    enum_out = tmp_path / "enum_out"
    _write_generated_enum(
        enum_out / "Quality.cs",
        "Quality",
        [("Common", 0), ("Rare", 5)],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    exported = {}

    def _fake_generate_enum_file(enum_type_name, enum_names, enum_values, remarks, name_space, output_folder):
        exported["enum_type_name"] = enum_type_name
        exported["enum_names"] = enum_names
        exported["enum_values"] = enum_values

    monkeypatch.setattr(export_process, "generate_enum_file", _fake_generate_enum_file)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=str(enum_out),
        auto_cleanup=False,
    )

    assert exported["enum_type_name"] == "Quality"
    assert exported["enum_names"] == ["Common", "Epic", "Rare"]
    assert exported["enum_values"] == [0, 6, 5]

    manifest = json.loads((tmp_path / ".stable_enum_values.json").read_text(encoding="utf-8"))
    assert manifest["Quality"]["Rare"] == 5


def test_batch_excel_to_json_all_explicit_enum_ignores_bootstrap_values(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", 10, "common"),
            ("Rare", 11, "rare"),
        ],
    )
    enum_out = tmp_path / "enum_out"
    _write_generated_enum(
        enum_out / "Quality.cs",
        "Quality",
        [("Common", 0), ("Rare", 5)],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)

    exported = {}

    def _fake_generate_enum_file(enum_type_name, enum_names, enum_values, remarks, name_space, output_folder):
        exported["enum_values"] = enum_values

    monkeypatch.setattr(export_process, "generate_enum_file", _fake_generate_enum_file)

    reset_enum_registry()
    export_process.batch_excel_to_json(
        source_folder=str(tmp_path),
        output_client_folder=None,
        output_project_folder=None,
        csfile_output_folder=None,
        enum_output_folder=str(enum_out),
        auto_cleanup=False,
    )

    assert exported["enum_values"] == [10, 11]
    manifest = json.loads((tmp_path / ".stable_enum_values.json").read_text(encoding="utf-8"))
    assert "Quality" not in manifest


def test_batch_excel_to_json_rejects_duplicate_manual_enum_values(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", 1, None),
            ("Rare", 1, None),
        ],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)
    enum_out = tmp_path / "enum_out"
    enum_out.mkdir(parents=True, exist_ok=True)

    reset_enum_registry()
    with pytest.raises(ExportError, match="duplicate enum value"):
        export_process.batch_excel_to_json(
            source_folder=str(tmp_path),
            output_client_folder=None,
            output_project_folder=None,
            csfile_output_folder=None,
            enum_output_folder=str(enum_out),
            auto_cleanup=False,
        )


def test_batch_excel_to_json_rejects_default_value_conflicting_with_explicit_value(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", None, None),
            ("Rare", 0, None),
        ],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)
    enum_out = tmp_path / "enum_out"
    enum_out.mkdir(parents=True, exist_ok=True)

    reset_enum_registry()
    with pytest.raises(ExportError, match="default enum value"):
        export_process.batch_excel_to_json(
            source_folder=str(tmp_path),
            output_client_folder=None,
            output_project_folder=None,
            csfile_output_folder=None,
            enum_output_folder=str(enum_out),
            auto_cleanup=False,
        )


def test_batch_excel_to_json_rejects_invalid_manual_enum_values(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", "1.5", None),
        ],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)
    enum_out = tmp_path / "enum_out"
    enum_out.mkdir(parents=True, exist_ok=True)

    reset_enum_registry()
    with pytest.raises(ExportError, match="invalid enum value"):
        export_process.batch_excel_to_json(
            source_folder=str(tmp_path),
            output_client_folder=None,
            output_project_folder=None,
            csfile_output_folder=None,
            enum_output_folder=str(enum_out),
            auto_cleanup=False,
        )


def test_batch_excel_to_json_rejects_explicit_value_conflicting_with_stable_default(monkeypatch, tmp_path):
    xlsx = tmp_path / "Item.xlsx"
    (tmp_path / ".stable_enum_values.json").write_text(
        json.dumps({"Quality": {"Common": 0, "Rare": 1}}, indent=2),
        encoding="utf-8",
    )
    _build_workbook_with_enum_sheet(
        xlsx,
        [
            ("Common", None, "common"),
            ("Epic", 1, "epic"),
            ("Rare", None, "rare"),
        ],
    )

    monkeypatch.setattr(export_process, "process_excel_file", lambda *args, **kwargs: None)
    enum_out = tmp_path / "enum_out"
    enum_out.mkdir(parents=True, exist_ok=True)

    reset_enum_registry()
    with pytest.raises(ExportError, match="conflicts with stable enum value"):
        export_process.batch_excel_to_json(
            source_folder=str(tmp_path),
            output_client_folder=None,
            output_project_folder=None,
            csfile_output_folder=None,
            enum_output_folder=str(enum_out),
            auto_cleanup=False,
        )
