# Author: huhongwei 306463233@qq.com
# MIT License
import time
import openpyxl
from pathlib import Path
from typing import Optional
import sys
import os
import shutil
import traceback
from .worksheet_data import WorksheetData
from .cs_generation import generate_enum_file_from_sheet, generate_enum_file, get_created_files, set_output_options
from .log import log_info, log_warn, log_error, log_success, log_sep, green_filename, log_warn_summary
from .exceptions import SheetNameConflictError, ExportError
from .exceptions import InvalidFieldNameError
from .exceptions import WriteFileError
from .exceptions import DuplicateFieldError, HeaderFormatError, UnknownCustomTypeError
from .enum_registry import get_enum_registry, reset_enum_registry
from .naming_config import ENUM_KEYS_SUFFIX
REPORT = None  # 报表文件输出已移除

def process_excel_file(
    excel_path: Path,
    file_sheet_map: dict[str, str],
    output_client_folder: Optional[str],
    output_project_folder: Optional[str],
    csfile_output_folder: Optional[str],
    enum_output_folder: Optional[str],
) -> Optional[WorksheetData]:
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    except Exception as e:
        log_error(f"打开失败: {green_filename(excel_path.name)} -> {e}")
        return None
    main_sheet = wb.worksheets[0]
    if main_sheet.title in file_sheet_map.values():
        dup = next(f for f, s in file_sheet_map.items() if s == main_sheet.title)
        raise SheetNameConflictError(main_sheet.title, dup, excel_path.name)

    log_sep(f"开始 {green_filename(excel_path.name)}")
    main_sheet_data = WorksheetData(main_sheet)
    # 记录来源 Excel 文件名，供日志使用
    try:
        setattr(main_sheet_data, "source_file", excel_path.name)
    except Exception:
        pass

    if output_client_folder:
        main_sheet_data.generate_json(output_client_folder)
    if output_project_folder:
        main_sheet_data.generate_json(output_project_folder)
    if csfile_output_folder:
        main_sheet_data.generate_script(csfile_output_folder)

    # 注意：枚举导出已在第一阶段完成，这里不再重复导出

    file_sheet_map[excel_path.name] = main_sheet.title
    log_info(f"完成 {excel_path.name} \n")
    return main_sheet_data

def cleanup_files(output_folders):
    created = set(get_created_files())
    from pathlib import Path
    stale = []
    for folder in filter(None, output_folders):
        p = Path(folder)
        if not p.exists():
            continue
        for f in p.rglob("*"):
            if f.is_file() and f.suffix != ".meta":
                if str(f.resolve()) not in created and str(f.with_suffix(f.suffix + ".meta").resolve()) not in created:
                    stale.append(f)
    if not stale:
        log_info("没有需要删除的文件")
        return
    # 临时切换 warning 为立即输出
    log_warn("以下文件未在本次生成中出现：", immediate=True)
    for f in stale:
        log_warn(f" - {f}", immediate=True)
    from .worksheet_data import user_confirm
    msg = "是否删除这些文件?(y/n): "
    if user_confirm(msg, title="文件删除确认"):
        for f in stale:
            try:
                f.unlink(missing_ok=True)
                log_info(f"删除: {f}")
            except Exception as e:
                log_error(f"删除失败 {f}: {e}")
    else:
        log_warn("已取消清理")

def batch_excel_to_json(
    source_folder: str,
    output_client_folder: Optional[str] = None,
    output_project_folder: Optional[str] = None,
    csfile_output_folder: Optional[str] = None,
    enum_output_folder: Optional[str] = None,
    diff_only: bool = True,
    dry_run: bool = False,
    auto_cleanup: bool = True,
) -> None:
    start = time.time()
    log_sep("开始导表")
    log_info(f"Excel目录: {source_folder}")
    set_output_options(diff_only=diff_only, dry_run=dry_run)

    # 输出目录可写性与磁盘空间检查
    def _check_output_dir(folder: Optional[str]) -> None:
        if not folder:
            return
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception as e:
            raise ExportError(f"无法创建输出目录: {folder} -> {e}")
        # 检查是否可写
        if not os.access(folder, os.W_OK):
            raise ExportError(f"输出目录不可写: {folder}")
        # 检查剩余空间（简单策略：至少 10MB 可用）
        try:
            stat = os.statvfs(folder)
            free = stat.f_bavail * stat.f_frsize
            if free < 10 * 1024 * 1024:
                raise ExportError(f"输出目录磁盘空间不足 (<10MB): {folder}")
        except AttributeError:
            # Windows 不支持 statvfs，尝试使用 shutil.disk_usage
            try:
                du = shutil.disk_usage(folder)
                if du.free < 10 * 1024 * 1024:
                    raise ExportError(f"输出目录磁盘空间不足 (<10MB): {folder}")
            except Exception:
                pass

    import shutil
    _check_output_dir(output_client_folder)
    _check_output_dir(output_project_folder)
    _check_output_dir(csfile_output_folder)
    _check_output_dir(enum_output_folder)

    skip = 0
    ok = 0
    file_sheet_map: dict[str, str] = {}
    # 反查 map: sheet 名 -> excel 文件名（用于日志显示目标 Excel 文件）
    sheet_to_file_map: dict[str, str] = {}
    excel_files = list(Path(source_folder).rglob("*.xlsx"))

    if not excel_files:
        log_warn("未找到 .xlsx 文件")

    # ========== 第一阶段：收集并导出所有枚举 ==========
    log_sep("第一阶段：收集并导出枚举")
    reset_enum_registry()  # 重置枚举注册表
    enum_registry = get_enum_registry()
    
    # 收集所有需要导出的枚举
    enum_files_to_export: list[Tuple[Path, str, list, list]] = []  # (excel_path, enum_type_name, enum_data, remarks)
    
    for excel_path in excel_files:
        if not excel_path.name[0].isupper():
            continue
        
        try:
            wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        except Exception as e:
            log_error(f"打开失败（枚举收集阶段）: {green_filename(excel_path.name)} -> {e}")
            continue
        
        main_sheet = wb.worksheets[0]
        
        # 方式1：检查主键是否为string类型（需要生成枚举）
        try:
            # 读取类型行（第3行，索引从1开始）
            if len(main_sheet[3]) > 0:
                first_field_type = str(main_sheet[3][0].value).strip().lower() if main_sheet[3][0].value else ""
                if first_field_type in ("str", "string"):
                    # 需要生成枚举，收集枚举项（从第7行开始）
                    enum_type_name = f"{main_sheet.title}{ENUM_KEYS_SUFFIX}"
                    enum_items = {}
                    enum_remarks = []  # 收集注释（来自第1行备注列）
                    idx_val = 0
                    for row in main_sheet.iter_rows(min_row=7, values_only=False):
                        if not row or not row[0]:
                            continue
                        val = row[0].value
                        if val is None:
                            continue
                        val_str = str(val).strip()
                        if val_str:
                            enum_items[val_str] = idx_val
                            # 收集注释：第1行（备注行）对应列的值
                            remark = None
                            if len(main_sheet[1]) > 0:  # 第1行存在
                                remark_cell = main_sheet[1][0]  # 第1行第1列（与数据行第1列对应）
                                if remark_cell and remark_cell.value:
                                    remark = str(remark_cell.value).strip()
                                    if not remark:
                                        remark = None
                            enum_remarks.append(remark)
                            idx_val += 1
                    
                    if enum_items:
                        # 验证枚举项名称格式
                        invalid_items = []
                        for item_name in enum_items.keys():
                            if not enum_registry.validate_enum_item_name(item_name):
                                invalid_items.append(item_name)
                        if invalid_items:
                            raise ExportError(
                                f"枚举 {enum_type_name} (来自 {excel_path.name}) 包含不符合C#命名规范的枚举项: {invalid_items}。"
                                f"枚举项必须以大写字母开头（大写驼峰式）"
                            )
                        enum_registry.register_enum(enum_type_name, enum_items, "Data.TableScript")
                        enum_files_to_export.append((excel_path, enum_type_name, list(enum_items.items()), enum_remarks))
                        log_info(f"收集枚举: {enum_type_name} (来自 {excel_path.name})")
        except Exception as e:
            log_warn(f"收集枚举时出错 {excel_path.name}: {e}")
        
        # 方式2：检查是否有Enum-开头的sheet
        if len(wb.worksheets) > 1 and enum_output_folder:
            enum_tag = "Enum-"
            for sheet in wb.worksheets[1:]:
                if sheet.title.startswith(enum_tag):
                    enum_type_name = sheet.title.replace(enum_tag, "")
                    enum_items = {}
                    enum_remarks = []  # 收集注释（来自第3列）
                    rows = list(sheet.iter_rows(min_row=2))
                    for r in rows:
                        if len(r) < 2:
                            continue
                        name = r[0].value
                        val = r[1].value
                        if name is None or val is None:
                            continue
                        name_str = str(name).strip()
                        try:
                            val_int = int(val)
                            enum_items[name_str] = val_int
                            # 收集注释：第3列（索引2）
                            remark = None
                            if len(r) > 2 and r[2].value:
                                remark = str(r[2].value).strip()
                                if not remark:
                                    remark = None
                            enum_remarks.append(remark)
                        except (ValueError, TypeError):
                            log_warn(f"{sheet.title} 枚举值非整数: {name}={val}")
                    
                    if enum_items:
                        # 验证枚举项名称格式
                        invalid_items = []
                        for item_name in enum_items.keys():
                            if not enum_registry.validate_enum_item_name(item_name):
                                invalid_items.append(item_name)
                        if invalid_items:
                            raise ExportError(
                                f"枚举 {enum_type_name} (来自 {excel_path.name}/{sheet.title}) 包含不符合C#命名规范的枚举项: {invalid_items}。"
                                f"枚举项必须以大写字母开头（大写驼峰式）"
                            )
                        enum_registry.register_enum(enum_type_name, enum_items, "Data.TableScript")
                        enum_files_to_export.append((excel_path, enum_type_name, list(enum_items.items()), enum_remarks))
                        log_info(f"收集枚举: {enum_type_name} (来自 {excel_path.name}/{sheet.title})")
        
        wb.close()
    
    # 导出所有枚举文件
    if enum_output_folder and enum_files_to_export:
        log_info(f"开始导出 {len(enum_files_to_export)} 个枚举...")
        for excel_path, enum_type_name, enum_data, remarks in enum_files_to_export:
            enum_names = [item[0] for item in enum_data]
            enum_values = [item[1] for item in enum_data]
            generate_enum_file(enum_type_name, enum_names, enum_values, remarks, "Data.TableScript", enum_output_folder)
            log_info(f"已导出枚举: {enum_type_name}")
    elif enum_output_folder:
        log_info("未发现需要导出的枚举")
    
    # ========== 第二阶段：处理表格数据 ==========
    log_sep("第二阶段：处理表格数据")
    sheets: list[WorksheetData] = []
    aborted = False
    for excel_path in excel_files:
        if not excel_path.name[0].isupper():
            log_warn(f"跳过(首字母非大写): {green_filename(excel_path.name)}")
            skip += 1
            continue
        try:
            ws = process_excel_file(
                excel_path,
                file_sheet_map,
                output_client_folder,
                output_project_folder,
                csfile_output_folder,
                enum_output_folder,
            )
            if ws is not None:
                sheets.append(ws)
                # 记录 sheet -> 文件名
                try:
                    sheet_to_file_map[ws.name] = excel_path.name
                except Exception:
                    pass
            ok += 1
        except Exception as e:
            # 所有异常视为致命：打印红色错误、堆栈信息，并给出建议后立即退出
            tb = traceback.format_exc()
            log_error(f"{excel_path.name} 失败: {e}\n{tb}")
            sys.exit(1)

    # 统一引用检查（导出后）
    if sheets and not aborted:
        search_dirs = [output_client_folder, output_project_folder]
        # 空行分隔阶段，并打印一次阶段标题
        log_info("")
        log_info("————开始引用检查————")
        for ws in sheets:
            try:
                ws.run_reference_checks(search_dirs, sheet_to_file_map)
            except Exception as e:
                log_error(f"[{ws.name}] 引用检查失败: {e}")

    # 打印每表错误/警告统计（若实现了内部统计则输出；当前由 worksheet 在控制台输出具体错误）

    if auto_cleanup and not aborted:
        log_sep("清理阶段")
        cleanup_files([output_client_folder, output_project_folder, csfile_output_folder, enum_output_folder])

    elapsed = time.time() - start
    log_sep("结束")
    # 统计本次生成的 JSON 文件总体积（仅统计已实际写入的文件）
    try:
        created_files = set(get_created_files())
        total_json_bytes = 0
        for p in created_files:
            try:
                if p.lower().endswith('.json') and os.path.isfile(p):
                    total_json_bytes += os.path.getsize(p)
            except Exception:
                # 忽略单个文件统计失败
                pass

        def _human_bytes(n: int) -> str:
            # 简单的人类可读格式
            if n < 1024:
                return f"{n} B"
            if n < 1024 * 1024:
                return f"{n/1024:.1f} KB"
            return f"{n/1024/1024:.2f} MB"

        total_json_str = _human_bytes(total_json_bytes)
    except Exception:
        total_json_bytes = 0
        total_json_str = "N/A"

    # 在打印最终结果前统一输出所有警告，便于快速查看
    try:
        log_warn_summary("以下为本次运行收集到的所有警告：")
    except Exception:
        pass
    if aborted:
        log_error(f"导表已中止: 字段命名不合法，已停止后续处理。成功 {ok}，跳过 {skip}，总耗时 {elapsed:.2f}s，总生成 JSON 大小: {total_json_str}")
    else:
        log_success(f"成功 {ok}，跳过 {skip}，总耗时 {elapsed:.2f}s，总生成 JSON 大小: {total_json_str}. diff_only:{diff_only}, dry_run:{dry_run}")
